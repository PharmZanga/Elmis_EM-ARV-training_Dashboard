import argparse
import csv
import json
import os
import re
from collections import Counter
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl


SOURCE_DIR = Path(r"C:\Users\Zanga Musakuzi\Desktop\ELMIS DASH BOARD\eLMIS Final Draft_20260226")
PROJECT_DIR = Path(__file__).resolve().parents[1]
OUT_FILE = PROJECT_DIR / "src" / "dashboardData.js"
RECON_DIR = PROJECT_DIR / "reconciliation"


def clean(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def normalized_code(value):
    text = clean(value).upper()
    text = re.sub(r"[^A-Z0-9]", "", text)
    # Excel sometimes turns numeric facility codes into values such as 101052.0.
    if text.endswith("0") and clean(value).endswith(".0"):
        text = text[:-1]
    return text.lstrip("0") or text


def normalized_name(value):
    text = clean(value).casefold()
    text = text.replace("centre", "center")
    text = re.sub(r"\brural health center\b", "rhc", text)
    text = re.sub(r"\bhealth center\b", "hc", text)
    text = re.sub(r"\bhealth post\b", "hp", text)
    text = re.sub(r"\bdistrict health office\b", "dho", text)
    return re.sub(r"[^a-z0-9]", "", text)


def period_label(value):
    if isinstance(value, datetime):
        return value.strftime("%B %Y")
    text = clean(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%B %Y")
        except ValueError:
            pass
    return text


def read_rows(path, sheet_name):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    headers = [clean(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
        if any(value is not None and clean(value) != "" for value in record.values()):
            yield record


def load_elmis_facility_metadata():
    try:
        import psycopg2
    except ImportError:
        print("eLMIS metadata: psycopg2 is not installed; using Excel facility fields.")
        return {}, {}, {}, {}

    connection = None
    try:
        connection = psycopg2.connect(
            host=os.getenv("ELMIS_DB_HOST", "127.0.0.1"),
            port=int(os.getenv("ELMIS_DB_PORT", "5432")),
            dbname=os.getenv("ELMIS_DB_NAME", "openlmis_local"),
            user=os.getenv("ELMIS_DB_USER", "postgres"),
            password=os.getenv("ELMIS_DB_PASSWORD", ""),
            connect_timeout=5,
        )
        with connection.cursor() as cursor:
            cursor.execute('''
                SELECT id, code, name, type, district_name, "provence" AS province
                FROM public.dim_facility
                WHERE code IS NOT NULL AND BTRIM(code) <> ''
                ORDER BY code
            ''')
            rows = cursor.fetchall()

        by_code, by_name, by_norm_code, by_norm_name = {}, {}, {}, {}
        for facility_id, code, name, facility_type, district, province in rows:
            item = {
                "id": facility_id,
                "code": clean(code),
                "name": clean(name),
                "type": clean(facility_type),
                "district": clean(district),
                "province": clean(province),
            }
            by_code[item["code"].upper()] = item
            if item["name"]:
                by_name.setdefault(item["name"].casefold(), item)
                by_norm_name.setdefault(normalized_name(item["name"]), item)
            by_norm_code.setdefault(normalized_code(item["code"]), item)

        print(f"eLMIS metadata: loaded {len(by_code)} facilities from public.dim_facility")
        return by_code, by_name, by_norm_code, by_norm_name
    except Exception as exc:
        print(f"eLMIS metadata: database unavailable ({exc}); using Excel facility fields.")
        return {}, {}, {}, {}
    finally:
        if connection is not None:
            connection.close()


FACILITIES_BY_CODE, FACILITIES_BY_NAME, FACILITIES_BY_NORM_CODE, FACILITIES_BY_NORM_NAME = load_elmis_facility_metadata()


def authoritative_facility(row):
    code = clean(row.get("Facility Code")).upper()
    if code and code in FACILITIES_BY_CODE:
        return FACILITIES_BY_CODE[code]
    name = clean(row.get("Facility"))
    if name:
        return FACILITIES_BY_NAME.get(name.casefold())
    return None


def participants():
    path = SOURCE_DIR / "Participants Masterfile.xlsx"
    sheets = {"eLMIS Expert ToT": "Expert", "eLMIS ToT Superusers": "Superuser", "eLMIS Trained Users": "User"}
    records = []
    for sheet, role in sheets.items():
        for row in read_rows(path, sheet):
            records.append({
                "role": role, "province": clean(row.get("Province")), "district": clean(row.get("District")),
                "firstName": clean(row.get("First Name")), "lastName": clean(row.get("Last Name")),
                "phone": clean(row.get("Mobile Phone")), "nrc": clean(row.get("NRC")),
                "facility": clean(row.get("Duty Station")), "profession": clean(row.get("Profession")).title() or "Not Specified",
                "startDate": clean(row.get("Start Date")), "endDate": clean(row.get("End Date")),
                "issuesResolved": int(row.get("Issues Resolved") or 0),
            })
    return records


def reporting_rows():
    path = SOURCE_DIR / "Reporting Status Masterfile.xlsx"
    records, matched, unmatched = [], 0, 0
    for row in read_rows(path, "Page 1"):
        meta = authoritative_facility(row)
        matched += bool(meta)
        unmatched += not bool(meta)
        records.append({
            "facilityCode": meta["code"] if meta else clean(row.get("Facility Code")),
            "facility": meta["name"] if meta else clean(row.get("Facility")),
            "facilityType": meta["type"] if meta else clean(row.get("Facility type")),
            "province": meta["province"] if meta else clean(row.get("Province")),
            "district": meta["district"] if meta else clean(row.get("District")),
            "program": clean(row.get("Program")), "period": period_label(row.get("Period")),
            "dateReceived": clean(row.get("Date Report Received")), "status": clean(row.get("Status")),
            "metadataMatched": bool(meta),
        })
    if FACILITIES_BY_CODE:
        print(f"Reporting rows matched to eLMIS metadata: {matched}; unmatched: {unmatched}")
    return records


def timeliness_rows():
    path = SOURCE_DIR / "Timeliness Reporting Masterfile.xlsx"
    records = []
    for row in read_rows(path, "Page 1"):
        expected, on_time, late = int(row.get("Expected") or 0), int(row.get("Reported On Time") or 0), int(row.get("Reported Late") or 0)
        records.append({
            "district": clean(row.get("District")), "province": clean(row.get("Region")),
            "supplyingDepot": clean(row.get("Supplying Depot")), "expected": expected,
            "reportedOnTime": on_time, "reportedLate": late, "period": period_label(row.get("Period")),
            "program": clean(row.get("Program")), "timeliness": round((on_time / expected) * 100, 2) if expected else 0,
        })
    return records


def facility_metadata():
    return sorted(FACILITIES_BY_CODE.values(), key=lambda item: (item["province"], item["district"], item["name"]))


def export_unmatched_report():
    source = SOURCE_DIR / "Reporting Status Masterfile.xlsx"
    grouped = {}
    for row in read_rows(source, "Page 1"):
        if authoritative_facility(row):
            continue
        key = (clean(row.get("Facility Code")), clean(row.get("Facility")), clean(row.get("District")), clean(row.get("Province")))
        item = grouped.setdefault(key, {"rows": 0, "programs": set(), "periods": set()})
        item["rows"] += 1
        item["programs"].add(clean(row.get("Program")))
        item["periods"].add(period_label(row.get("Period")))

    RECON_DIR.mkdir(exist_ok=True)
    out = RECON_DIR / "unmatched_facilities.csv"
    categories = Counter()
    fields = ["facility_code", "facility_name", "district", "province", "reporting_rows", "programs", "periods", "reconciliation_category", "suggested_elmis_code", "suggested_elmis_name", "suggested_district", "suggested_province", "name_similarity"]

    with out.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for (code, name, district, province), detail in sorted(grouped.items(), key=lambda x: (-x[1]["rows"], x[0][1])):
            suggestion = None
            category = "NOT_IN_METADATA"
            score = 0.0

            ncode = normalized_code(code)
            if ncode and ncode in FACILITIES_BY_NORM_CODE:
                suggestion = FACILITIES_BY_NORM_CODE[ncode]
                category = "CODE_FORMAT_DIFFERENCE"
                score = 1.0
            else:
                nname = normalized_name(name)
                if nname and nname in FACILITIES_BY_NORM_NAME:
                    suggestion = FACILITIES_BY_NORM_NAME[nname]
                    category = "NAME_FORMAT_DIFFERENCE"
                    score = 1.0
                elif nname:
                    candidates = list(FACILITIES_BY_CODE.values())
                    same_district = [x for x in candidates if clean(x["district"]).casefold() == district.casefold()]
                    if same_district:
                        candidates = same_district
                    best = None
                    for candidate in candidates:
                        similarity = SequenceMatcher(None, nname, normalized_name(candidate["name"])).ratio()
                        if best is None or similarity > best[0]:
                            best = (similarity, candidate)
                    if best and best[0] >= 0.86:
                        score, suggestion = best
                        category = "POSSIBLE_NAME_MATCH"

            categories[category] += 1
            writer.writerow({
                "facility_code": code, "facility_name": name, "district": district, "province": province,
                "reporting_rows": detail["rows"], "programs": " | ".join(sorted(x for x in detail["programs"] if x)),
                "periods": " | ".join(sorted(x for x in detail["periods"] if x)), "reconciliation_category": category,
                "suggested_elmis_code": suggestion["code"] if suggestion else "", "suggested_elmis_name": suggestion["name"] if suggestion else "",
                "suggested_district": suggestion["district"] if suggestion else "", "suggested_province": suggestion["province"] if suggestion else "",
                "name_similarity": f"{score:.3f}" if score else "",
            })

    print(f"Unmatched reconciliation report: {out}")
    print(f"Unique unmatched facility records: {len(grouped)}")
    print("Reconciliation categories: " + ", ".join(f"{k}={v}" for k, v in sorted(categories.items())))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--export-unmatched", action="store_true", help="Export unique unmatched facilities with reconciliation suggestions")
    args = parser.parse_args()

    data = {
        "participants": participants(), "reportingRows": reporting_rows(),
        "timelinessRows": timeliness_rows(), "facilityMetadata": facility_metadata(),
    }
    OUT_FILE.write_text("export const dashboardData = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    print(f"Wrote {OUT_FILE}")
    print(f"Participants: {len(data['participants'])}")
    print(f"Reporting rows: {len(data['reportingRows'])}")
    print(f"Timeliness rows: {len(data['timelinessRows'])}")
    print(f"Facility metadata rows: {len(data['facilityMetadata'])}")
    if args.export_unmatched:
        export_unmatched_report()


if __name__ == "__main__":
    main()
